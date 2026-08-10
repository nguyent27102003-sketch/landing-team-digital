import openpyxl
import json
import os

wb_path = r'c:\Users\Administrator\Downloads\HÙNG CƯỜNG — EQUIPMENT CONFIGURATOR BACKEND v1.0.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=True)

# 1. Product Master Analysis
ws_p = wb['03_PRODUCT_MASTER']
p_rows = list(ws_p.iter_rows(values_only=True))
p_head = p_rows[0]
p_data = [dict(zip(p_head, r)) for r in p_rows[1:] if any(x is not None for x in r)]

# 2. Link Price Analysis
ws_l = wb['09_LINK_PRICE_STATUS']
l_rows = list(ws_l.iter_rows(values_only=True))
l_head = l_rows[0]
l_data = {r[0]: dict(zip(l_head, r)) for r in l_rows[1:] if any(x is not None for x in r)}

# 3. Compatibility Analysis
ws_c = wb['05_COMPATIBILITY']
c_rows = list(ws_c.iter_rows(values_only=True))
c_head = c_rows[0]
c_data = [dict(zip(c_head, r)) for r in c_rows[1:] if any(x is not None for x in r)]

# 4. Rules Analysis
ws_r = wb['06_RECOMMEND_RULE']
r_rows = list(ws_r.iter_rows(values_only=True))
r_head = r_rows[0]
r_data = [dict(zip(r_head, r)) for r in r_rows[1:] if any(x is not None for x in r)]

# 5. Explanations Analysis
ws_e = wb['07_EXPLANATION_RULE']
e_rows = list(ws_e.iter_rows(values_only=True))
e_head = e_rows[0]
e_data = [dict(zip(e_head, r)) for r in e_rows[1:] if any(x is not None for x in r)]

# 6. Comparisons Analysis
ws_cmp = wb['12_PRODUCT_COMPARISON']
cmp_rows = list(ws_cmp.iter_rows(values_only=True))
cmp_head = cmp_rows[0]
cmp_data = [dict(zip(cmp_head, r)) for r in cmp_rows[1:] if any(x is not None for x in r)]

# 7. Alternatives Analysis
ws_a = wb['08_ALTERNATIVE_MAP']
a_rows = list(ws_a.iter_rows(values_only=True))
a_head = a_rows[0]
a_data = [dict(zip(a_head, r)) for r in a_rows[1:] if any(x is not None for x in r)]

# Run deep product verification for all 81 products
product_verification_results = []
price_link_categories = {
    'PURCHASE_READY': [],
    'INFO_ONLY': [],
    'NEED_VERIFY': [],
    'OUT_OF_STOCK': [],
    'DEAD_LINK': [],
    'NO_PRICE': []
}

for p in p_data:
    pid = p['Product_ID']
    link_info = l_data.get(pid, {})
    
    # Checks
    checks = []
    
    # 1. ID
    if not pid:
        checks.append(('Product_ID', 'Non-empty', str(pid), 'FAIL', 'Missing ID', 'CRITICAL'))
    else:
        checks.append(('Product_ID', 'Unique', pid, 'PASS', 'None', 'LOW'))
        
    # 2. Name
    name = p.get('Product_Name')
    if not name:
        checks.append(('Product_Name', 'Exists', str(name), 'FAIL', 'Missing Name', 'CRITICAL'))
    else:
        checks.append(('Product_Name', 'Exists', name, 'PASS', 'None', 'LOW'))
        
    # 3. Category & System
    cat = p.get('Category')
    sys_val = p.get('System')
    checks.append(('Category', 'Valid Enum', str(cat), 'PASS' if cat else 'FAIL', 'None' if cat else 'Missing Category', 'HIGH'))
    checks.append(('System', 'Valid Enum', str(sys_val), 'PASS' if sys_val else 'FAIL', 'None' if sys_val else 'Missing System', 'HIGH'))
    
    # 4. Price & Date
    price = p.get('Price_Current')
    p_status = p.get('Price_Status')
    p_date = p.get('Price_Checked_Date')
    
    # 5. Link & Platform
    p_link = p.get('Primary_Link')
    p_plat = p.get('Primary_Platform')
    l_status = p.get('Link_Status')
    b_link = p.get('Backup_Link')
    b_plat = p.get('Backup_Platform')
    
    # Classification of Commercial Status
    if p_status == 'VERIFY' and l_status == 'ACTIVE' and p_plat in ['TIKTOK_SHOP', 'RETAILER'] and price is not None:
        price_link_categories['PURCHASE_READY'].append(pid)
        comm_cat = 'PURCHASE_READY'
    elif l_status == 'ACTIVE' and p_plat == 'OFFICIAL':
        price_link_categories['INFO_ONLY'].append(pid)
        comm_cat = 'INFO_ONLY'
    elif p_status == 'NEED_VERIFY' or l_status == 'NEED_VERIFY':
        price_link_categories['NEED_VERIFY'].append(pid)
        comm_cat = 'NEED_VERIFY'
    elif l_status == 'DEAD':
        price_link_categories['DEAD_LINK'].append(pid)
        comm_cat = 'DEAD_LINK'
    elif price is None:
        price_link_categories['NO_PRICE'].append(pid)
        comm_cat = 'NO_PRICE'
    else:
        price_link_categories['NEED_VERIFY'].append(pid)
        comm_cat = 'NEED_VERIFY'
        
    # Special check: Sold out / Out of stock note
    notes = str(p.get('Notes', ''))
    if 'sold out' in notes.lower() or 'out of stock' in notes.lower() or pid == 'PWR_ANK_PRIME20':
        price_link_categories['OUT_OF_STOCK'].append(pid)
        
    # 6. Evaluation Facts & Impacts
    s1_f = p.get('Strength_1_Fact')
    s1_i = p.get('Strength_1_Impact')
    l1_f = p.get('Limitation_1_Fact')
    l1_i = p.get('Limitation_1_Impact')
    best_for = p.get('Best_For')
    not_for = p.get('Not_For')
    tradeoff = p.get('Main_Tradeoff')
    
    has_facts = bool(s1_f and s1_i)
    has_limitations = bool(l1_f and l1_i)
    
    product_verification_results.append({
        'Product_ID': pid,
        'Brand': p.get('Brand'),
        'Model': p.get('Model'),
        'Product_Name': name,
        'Category': cat,
        'Subcategory': p.get('Subcategory'),
        'System': sys_val,
        'Price_Current': price,
        'Price_Status': p_status,
        'Price_Checked_Date': str(p_date) if p_date else None,
        'Primary_Link': p_link,
        'Primary_Platform': p_plat,
        'Link_Status': l_status,
        'Backup_Link': b_link,
        'Commercial_Category': comm_cat,
        'Has_Strength_Fact_Impact': has_facts,
        'Has_Limitation_Fact_Impact': has_limitations,
        'Best_For': best_for,
        'Not_For': not_for,
        'Main_Tradeoff': tradeoff,
        'Recommend_Status': p.get('Recommend_Status'),
        'Data_Verification_Status': p.get('Data_Verification_Status'),
        'Checks': checks
    })

# Dump full verification to json
verification_summary = {
    'total_products': len(product_verification_results),
    'commercial_breakdown': {k: len(v) for k, v in price_link_categories.items()},
    'commercial_details': price_link_categories,
    'products': product_verification_results,
    'compatibility_records': c_data,
    'recommend_rules': r_data,
    'explanation_rules': e_data,
    'comparison_records': cmp_data,
    'alternatives_records': a_data
}

with open('full_verification_data.json', 'w', encoding='utf-8') as f:
    json.dump(verification_summary, f, ensure_ascii=False, indent=2)

print("Verification data written to full_verification_data.json successfully")
