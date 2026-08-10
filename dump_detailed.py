import openpyxl, json

wb_path = r'c:\Users\Administrator\Downloads\HÙNG CƯỜNG — EQUIPMENT CONFIGURATOR BACKEND v1.0.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=True)
wb_formula = openpyxl.load_workbook(wb_path, data_only=False)

def get_sheet_details(sheet_name):
    ws = wb[sheet_name]
    ws_f = wb_formula[sheet_name]
    
    rows = list(ws.iter_rows(values_only=True))
    rows_f = list(ws_f.iter_rows(values_only=True))
    
    header = rows[0] if rows else []
    # Clean header
    header = [str(h) if h is not None else f"Col_{i+1}" for i, h in enumerate(header)]
    
    data = []
    for r_idx in range(1, len(rows)):
        r = rows[r_idx]
        rf = rows_f[r_idx]
        if not any(x is not None for x in r):
            continue
        row_dict = {}
        formulas = {}
        for c_idx in range(len(r)):
            col_name = header[c_idx] if c_idx < len(header) else f"Col_{c_idx+1}"
            val = r[c_idx]
            f_val = rf[c_idx]
            row_dict[col_name] = val
            if str(f_val).startswith('='):
                formulas[col_name] = str(f_val)
        data.append({
            'row_index': r_idx + 1,
            'values': row_dict,
            'formulas': formulas
        })
    return {
        'header': header,
        'count': len(data),
        'data': data
    }

all_sheets = {}
for name in wb.sheetnames:
    all_sheets[name] = get_sheet_details(name)

with open('detailed_sheets.json', 'w', encoding='utf-8') as f:
    json.dump(all_sheets, f, ensure_ascii=False, indent=2, default=str)

print("detailed_sheets.json written successfully.")
