import openpyxl
import json
import os

wb_path = r'c:\Users\Administrator\Downloads\HÙNG CƯỜNG — EQUIPMENT CONFIGURATOR BACKEND v1.0.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=True)
wb_formula = openpyxl.load_workbook(wb_path, data_only=False)

analysis = {}

for name in wb.sheetnames:
    ws = wb[name]
    ws_f = wb_formula[name]
    
    rows = list(ws.iter_rows(values_only=True))
    rows_f = list(ws_f.iter_rows(values_only=True))
    
    # Filter out empty rows
    non_empty_indices = [i for i, r in enumerate(rows) if any(x is not None for x in r)]
    
    analysis[name] = {
        'total_rows': len(rows),
        'non_empty_count': len(non_empty_indices),
        'rows': [list(rows[i]) for i in non_empty_indices],
        'formulas_sample': [
            {f"R{i+1}C{j+1}": str(rows_f[i][j]) for j in range(len(rows_f[i])) if str(rows_f[i][j]).startswith('=')}
            for i in non_empty_indices[:15]
            if any(str(rows_f[i][j]).startswith('=') for j in range(len(rows_f[i])))
        ]
    }

with open('full_workbook_dump.json', 'w', encoding='utf-8') as f:
    json.dump(analysis, f, ensure_ascii=False, indent=2, default=str)

# Also create a readable markdown audit report
md_lines = []
md_lines.append("# WORKBOOK AUDIT DETAILS\n")

for name, data in analysis.items():
    md_lines.append(f"\n## Sheet: `{name}` (Non-empty rows: {data['non_empty_count']})")
    rows = data['rows']
    if not rows:
        md_lines.append("*(Empty sheet)*")
        continue
    
    # Print header and first few rows
    md_lines.append("\n### Sample Data:")
    for idx, r in enumerate(rows[:10]):
        cleaned = [str(x).strip().replace('\n', ' ') if x is not None else '' for x in r]
        # remove trailing empty
        while cleaned and cleaned[-1] == '':
            cleaned.pop()
        md_lines.append(f"- **Row {idx+1}**: `{' | '.join(cleaned[:15])}`")
    
    if data['formulas_sample']:
        md_lines.append("\n### Sample Formulas:")
        for f_item in data['formulas_sample'][:5]:
            md_lines.append(f"- `{f_item}`")

with open('workbook_audit_report.md', 'w', encoding='utf-8') as f:
    f.write('\n'.join(md_lines))

print("Audit files written successfully: full_workbook_dump.json, workbook_audit_report.md")
